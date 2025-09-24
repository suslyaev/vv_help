from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Count
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.urls import reverse
from django.utils.safestring import mark_safe
from .models import Ticket, Category, Client, Organization, TicketStatus, TicketComment, TicketTemplate, TicketAudit, TicketAttachment, TelegramMessage, TelegramRoute, TelegramGroup
from .forms import TicketForm, TicketCommentForm, ClientForm, TicketAttachmentForm, OrganizationForm


@login_required
def dashboard(request):
    """Главная страница с дашбордом"""
    # Статистика
    total_tickets = Ticket.objects.count()
    open_tickets = Ticket.objects.filter(status__is_final=False).count()
    my_tickets = Ticket.objects.filter(assigned_to=request.user, status__is_final=False).count()
    overdue_tickets = Ticket.objects.filter(
        status__is_final=False,
        created_at__lt=timezone.now() - timezone.timedelta(hours=24)
    ).count()
    
    # Последние обращения
    recent_tickets = Ticket.objects.select_related(
        'client', 'category', 'status', 'assigned_to'
    ).order_by('-created_at')[:10]
    
    # Обращения по статусам
    status_stats = TicketStatus.objects.annotate(
        ticket_count=Count('ticket')
    ).exclude(name='Закрыто').order_by('order')
    
    # Обращения по категориям
    category_stats = Category.objects.annotate(
        ticket_count=Count('ticket')
    ).filter(ticket_count__gt=0).order_by('-ticket_count')[:10]
    
    context = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'my_tickets': my_tickets,
        'overdue_tickets': overdue_tickets,
        'recent_tickets': recent_tickets,
        'status_stats': status_stats,
        'category_stats': category_stats,
    }
    
    return render(request, 'tickets/dashboard.html', context)


@login_required
def ticket_list(request):
    """Список обращений с фильтрацией"""
    tickets = Ticket.objects.select_related(
        'client', 'organization', 'category', 'status', 'assigned_to'
    ).order_by('-created_at')
    
    # Фильтры
    status_filter = request.GET.get('status')
    # Поддержка нового autocomplete: приходят category_id и текст category
    category_filter = request.GET.get('category_id') or request.GET.get('category')
    if category_filter in (None, '', 'None', 'null', 'NULL'):
        category_filter = None
    assigned_filter = request.GET.get('assigned')
    organization_filter = request.GET.get('organization_id') or request.GET.get('organization')
    if organization_filter in (None, '', 'None', 'null', 'NULL'):
        organization_filter = None
    search_query = request.GET.get('search')
    
    if status_filter:
        tickets = tickets.filter(status_id=status_filter)
    
    if category_filter:
        try:
            tickets = tickets.filter(category_id=int(category_filter))
        except (TypeError, ValueError):
            pass
    
    if assigned_filter == 'me':
        tickets = tickets.filter(assigned_to=request.user)
    elif assigned_filter == 'unassigned':
        tickets = tickets.filter(assigned_to__isnull=True)
    
    if organization_filter:
        try:
            tickets = tickets.filter(organization_id=int(organization_filter))
        except (TypeError, ValueError):
            pass
    
    if search_query:
        tickets = tickets.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(client__name__icontains=search_query) |
            Q(tags__icontains=search_query)
        )
    
    # Пагинация
    paginator = Paginator(tickets, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Данные для фильтров (категории больше не нужны для select)
    statuses = TicketStatus.objects.exclude(name='Закрыто').order_by('order')
    
    context = {
        'page_obj': page_obj,
        # 'categories': categories,  # не требуется из-за autocomplete
        'statuses': statuses,
        'current_filters': {
            'status': status_filter,
            'category': category_filter,
            'assigned': assigned_filter,
            'organization': organization_filter,
            'search': search_query,
        }
    }
    
    return render(request, 'tickets/ticket_list.html', context)


@login_required
def ticket_detail(request, ticket_id):
    """Детальная страница обращения"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    comments = ticket.comments.select_related('author', 'author_client').order_by('created_at')
    attachments = ticket.attachments.select_related('uploaded_by').order_by('-uploaded_at')
    
    # Инициализируем формы
    comment_form = TicketCommentForm()
    attachment_form = TicketAttachmentForm()
    
    if request.method == 'POST':
        # Проверяем, какой тип формы отправлен
        if 'comment' in request.POST:
            comment_form = TicketCommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.ticket = ticket
                
                # Устанавливаем автора в зависимости от типа
                if comment.author_type == 'client':
                    # Получаем ID клиента из скрытого поля автокомплита
                    # Основное имя скрытого поля соответствует имени текстового поля + '_id'
                    # Для поля author_client_text это будет author_client_text_id
                    client_id = (
                        request.POST.get('author_client_text_id')
                        or request.POST.get('author_client_id')
                    )
                    if client_id and client_id.isdigit():
                        try:
                            comment.author_client = Client.objects.get(id=int(client_id))
                            comment.author = None  # Очищаем автора-пользователя
                        except Client.DoesNotExist:
                            pass
                else:  # author_type == 'user'
                    comment.author = request.user
                    comment.author_client = None  # Очищаем автора-клиента
                
                # Клиенты не могут создавать внутренние комментарии
                if comment.author_type == 'client':
                    comment.is_internal = False
                
                comment.save()
                
                # Проверяем, нужно ли отправить комментарий в Telegram
                # Отправляем в Telegram только если автор - системный пользователь
                reply_in_chat = request.POST.get('reply_in_chat') == '1' and comment.author_type == 'user'
                if reply_in_chat and ticket.telegram_chat_id and ticket.external_message_id:
                    try:
                        import logging
                        import asyncio
                        from telegram.ext import Application
                        from django.conf import settings
                        
                        logger = logging.getLogger(__name__)
                        logger.info(f"Attempting to send Telegram comment from ticket detail: chat_id={ticket.telegram_chat_id}, message_id={ticket.external_message_id}")
                        
                        bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                        if bot_token:
                            # Создаем асинхронную функцию для отправки комментария
                            async def send_telegram_comment():
                                application = Application.builder().token(bot_token).build()
                                result = await application.bot.send_message(
                                    chat_id=ticket.telegram_chat_id,
                                    text=comment.content,
                                    reply_to_message_id=int(ticket.external_message_id)
                                )
                                return result
                            
                            # Запускаем асинхронную функцию
                            result = asyncio.run(send_telegram_comment())
                            logger.info(f"Telegram comment sent successfully: {result.message_id}")
                            
                            # Сохраняем ID сообщения в комментарии
                            comment.telegram_message_id = str(result.message_id)
                            comment.save()
                            
                            # Добавляем отправленное сообщение в поток
                            try:
                                TelegramMessage.objects.create(
                                    message_id=str(result.message_id),
                                    chat_id=str(ticket.telegram_chat_id),
                                    chat_title=ticket.telegram_chat_title or '',
                                    from_user_id=str(request.user.id),
                                    from_username=request.user.username,
                                    from_fullname=request.user.get_full_name() or request.user.username,
                                    text=comment.content,
                                    message_date=timezone.now(),
                                    created_at=timezone.now(),
                                    linked_ticket=ticket,
                                    linked_action='add_comment',
                                    reply_to_message_id=str(ticket.external_message_id)
                                )
                                logger.info(f"Added comment message to stream: {result.message_id}")
                            except Exception as stream_error:
                                logger.error(f"Failed to add comment message to stream: {stream_error}", exc_info=True)
                            
                            messages.success(request, 'Комментарий добавлен и отправлен в Telegram')
                        else:
                            logger.error("TELEGRAM_BOT_TOKEN not configured")
                            messages.success(request, 'Комментарий добавлен (Telegram бот не настроен)')
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Failed to send Telegram comment: {e}", exc_info=True)
                        messages.success(request, 'Комментарий добавлен (не удалось отправить в Telegram)')
                else:
                    messages.success(request, 'Комментарий добавлен')
                
                # Создаем запись аудита
                TicketAudit.objects.create(
                    ticket=ticket,
                    action='comment_added',
                    user=request.user,
                    comment=f'Добавлен комментарий: {comment.content[:50]}...'
                )
                
                return redirect('tickets:ticket_detail', ticket_id=ticket.id)
            else:
                # Форма невалидна, показываем ошибки
                pass
        elif request.POST.get('action') == 'reply_to_comment':
            # Обработка ответа на комментарий
            comment_id = request.POST.get('comment_id')
            reply_content = request.POST.get('reply_content', '')
            is_internal_reply = request.POST.get('is_internal_reply') == 'on'
            send_to_telegram = request.POST.get('send_to_telegram') == 'on'
            
            if not comment_id or not comment_id.isdigit():
                messages.error(request, 'Неверный ID комментария')
                return redirect('tickets:ticket_detail', ticket_id=ticket.id)
            
            if not reply_content.strip():
                messages.error(request, 'Введите текст ответа')
                return redirect('tickets:ticket_detail', ticket_id=ticket.id)
            
            try:
                original_comment = TicketComment.objects.get(id=int(comment_id), ticket=ticket)
            except TicketComment.DoesNotExist:
                messages.error(request, 'Комментарий не найден')
                return redirect('tickets:ticket_detail', ticket_id=ticket.id)
            
            # Создаем новый комментарий
            reply_comment = TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                author_type='user',
                author_client=None,
                content=reply_content,
                is_internal=is_internal_reply,
                created_at=timezone.now()
            )
            
            # Если запрошена отправка в Telegram
            if send_to_telegram and original_comment.telegram_message_id and ticket.telegram_chat_id:
                try:
                    import logging
                    import asyncio
                    from telegram.ext import Application
                    from django.conf import settings
                    
                    logger = logging.getLogger(__name__)
                    logger.info(f"Attempting to send Telegram reply: chat_id={ticket.telegram_chat_id}, reply_to_message_id={original_comment.telegram_message_id}")
                    
                    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                    if bot_token:
                        # Создаем асинхронную функцию для отправки ответа
                        async def send_telegram_reply():
                            application = Application.builder().token(bot_token).build()
                            result = await application.bot.send_message(
                                chat_id=ticket.telegram_chat_id,
                                text=reply_content,
                                reply_to_message_id=int(original_comment.telegram_message_id)
                            )
                            return result
                        
                        # Запускаем асинхронную функцию
                        result = asyncio.run(send_telegram_reply())
                        logger.info(f"Telegram reply sent successfully: {result.message_id}")
                        
                        # Сохраняем ID сообщения в комментарии
                        reply_comment.telegram_message_id = str(result.message_id)
                        reply_comment.save()
                        
                        # Добавляем отправленное сообщение в поток
                        try:
                            TelegramMessage.objects.create(
                                message_id=str(result.message_id),
                                chat_id=str(ticket.telegram_chat_id),
                                chat_title=ticket.telegram_chat_title or '',
                                from_user_id=str(request.user.id),
                                from_username=request.user.username,
                                from_fullname=request.user.get_full_name() or request.user.username,
                                text=reply_content,
                                message_date=timezone.now(),
                                created_at=timezone.now(),
                                linked_ticket=ticket,
                                linked_action='add_comment',
                                reply_to_message_id=str(original_comment.telegram_message_id)
                            )
                            logger.info(f"Added reply message to stream: {result.message_id}")
                        except Exception as stream_error:
                            logger.error(f"Failed to add reply message to stream: {stream_error}", exc_info=True)
                        
                        messages.success(request, 'Ответ добавлен и отправлен в Telegram')
                    else:
                        logger.error("TELEGRAM_BOT_TOKEN not configured")
                        messages.success(request, 'Ответ добавлен (Telegram бот не настроен)')
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to send Telegram reply: {e}", exc_info=True)
                    messages.success(request, 'Ответ добавлен (не удалось отправить в Telegram)')
            else:
                messages.success(request, 'Ответ добавлен')
            
            # Создаем запись аудита
            TicketAudit.objects.create(
                ticket=ticket,
                action='comment_reply_added',
                user=request.user,
                comment=f'Добавлен ответ на комментарий: {reply_content[:50]}...'
            )
            
            return redirect('tickets:ticket_detail', ticket_id=ticket.id)
        elif 'attachment' in request.FILES:
            form = TicketAttachmentForm(request.POST, request.FILES)
            if form.is_valid():
                files = request.FILES.getlist('file')
                uploaded_count = 0
                
                for file in files:
                    attachment = TicketAttachment(
                        ticket=ticket,
                        file=file,
                        filename=file.name,
                        file_size=file.size,
                        uploaded_by=request.user
                    )
                    attachment.save()
                    uploaded_count += 1
                
                # Создаем запись аудита
                TicketAudit.objects.create(
                    ticket=ticket,
                    action='updated',
                    user=request.user,
                    comment=f'Загружено {uploaded_count} файлов'
                )
                
                messages.success(request, f'Загружено {uploaded_count} файлов')
                return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, можно ли отправлять комментарии в Telegram
    can_reply_in_telegram = False
    if ticket.telegram_chat_id and ticket.external_message_id:
        try:
            telegram_group = TelegramGroup.objects.get(chat_id=ticket.telegram_chat_id)
            can_reply_in_telegram = not telegram_group.is_blocked
        except TelegramGroup.DoesNotExist:
            can_reply_in_telegram = False
    
    # Обогащаем комментарии дополнительной информацией
    enriched_comments = []
    
    for comment in comments:
        comment_data = {
            'comment': comment,
            'is_reply': False,
            'reply_info': None,
            'from_bot': False
        }
        
        if comment.telegram_message_id:
            # Проверяем, является ли комментарий ответом
            telegram_msg = TelegramMessage.objects.filter(
                message_id=comment.telegram_message_id,
                chat_id=ticket.telegram_chat_id
            ).first()
            
            if telegram_msg and telegram_msg.reply_to_message_id:
                # Ищем исходное сообщение для отображения цитаты
                original_msg = TelegramMessage.objects.filter(
                    message_id=telegram_msg.reply_to_message_id,
                    chat_id=ticket.telegram_chat_id
                ).first()
                if original_msg:
                    comment_data['is_reply'] = True
                    comment_data['reply_info'] = {
                        'original_text': original_msg.text,
                        'original_author': original_msg.from_fullname or original_msg.from_username or 'Неизвестный'
                    }
            
            # Проверяем, отправлен ли комментарий ботом (от пользователя системы)
            from_bot = False
            if comment.author_type == 'user' and comment.telegram_message_id and comment.author:
                # Дополнительно проверяем, что в Telegram сообщении from_user_id соответствует текущему пользователю
                telegram_msg = TelegramMessage.objects.filter(
                    message_id=comment.telegram_message_id,
                    chat_id=ticket.telegram_chat_id
                ).first()
                
                if telegram_msg and telegram_msg.from_user_id == str(request.user.id):
                    from_bot = True
            
            comment_data['from_bot'] = from_bot
        
        enriched_comments.append(comment_data)
    
    context = {
        'ticket': ticket,
        'comments': comments,
        'enriched_comments': enriched_comments,
        'attachments': attachments,
        'comment_form': comment_form,
        'attachment_form': attachment_form,
        'can_reply_in_telegram': can_reply_in_telegram,
    }
    
    return render(request, 'tickets/ticket_detail.html', context)


@login_required
def edit_comment(request, comment_id):
    """Редактировать комментарий"""
    comment = get_object_or_404(TicketComment, id=comment_id)
    ticket = comment.ticket
    
    # Проверяем, что комментарий был отправлен в Telegram
    if not comment.telegram_message_id:
        messages.error(request, 'Редактирование недоступно для данного комментария')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что обращение имеет Telegram данные
    if not ticket.telegram_chat_id or not ticket.external_message_id:
        messages.error(request, 'Редактирование недоступно для данного обращения')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что группа разрешена для отправки ответов
    can_reply_in_telegram = False
    try:
        telegram_group = TelegramGroup.objects.get(chat_id=ticket.telegram_chat_id)
        can_reply_in_telegram = not telegram_group.is_blocked
    except TelegramGroup.DoesNotExist:
        can_reply_in_telegram = False
    
    if not can_reply_in_telegram:
        messages.error(request, 'Редактирование недоступно для данной группы')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    if request.method == 'POST':
        new_content = request.POST.get('content', '')
        update_in_chat = request.POST.get('update_in_chat') == '1'
        
        if not new_content:
            messages.error(request, 'Содержание комментария не может быть пустым')
            return redirect('tickets:edit_comment', comment_id=comment.id)
        
        # Обновляем комментарий
        comment.content = new_content
        comment.save()
        
        # Создаем запись аудита
        TicketAudit.objects.create(
            ticket=ticket,
            action='comment_edited',
            user=request.user,
            comment=f'Комментарий отредактирован: {new_content[:50]}...'
        )
        
        # Если запрошено обновление в Telegram
        if update_in_chat:
            try:
                import logging
                import asyncio
                from telegram.ext import Application
                from django.conf import settings
                
                logger = logging.getLogger(__name__)
                logger.info(f"Attempting to edit Telegram comment: chat_id={ticket.telegram_chat_id}, message_id={comment.telegram_message_id}")
                
                bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                if bot_token:
                    # Создаем асинхронную функцию для редактирования сообщения
                    async def edit_telegram_comment():
                        application = Application.builder().token(bot_token).build()
                        try:
                            # Пытаемся отредактировать существующее сообщение
                            result = await application.bot.edit_message_text(
                                chat_id=ticket.telegram_chat_id,
                                message_id=int(comment.telegram_message_id),
                                text=new_content
                            )
                            return result, 'edited'
                        except Exception as edit_error:
                            # Если не удалось отредактировать, отправляем новое сообщение как ответ
                            logger.warning(f"Could not edit comment, sending new one: {edit_error}")
                            result = await application.bot.send_message(
                                chat_id=ticket.telegram_chat_id,
                                text=new_content,
                                reply_to_message_id=int(ticket.external_message_id)
                            )
                            return result, 'new'
                    
                    # Запускаем асинхронную функцию
                    result, action_type = asyncio.run(edit_telegram_comment())
                    logger.info(f"Telegram comment {action_type} successfully: {result.message_id}")
                    
                    # Обновляем сообщение в потоке
                    try:
                        if action_type == 'edited':
                            # Обновляем существующее сообщение
                            telegram_message = TelegramMessage.objects.filter(
                                chat_id=ticket.telegram_chat_id,
                                message_id=comment.telegram_message_id,
                                linked_ticket=ticket,
                                linked_action='add_comment'
                            ).first()
                            
                            if telegram_message:
                                telegram_message.text = new_content
                                telegram_message.save()
                                logger.info(f"Updated comment message in stream: {telegram_message.id}")
                        else:
                            # Обновляем ID сообщения в комментарии
                            comment.telegram_message_id = str(result.message_id)
                            comment.save()
                            
                            # Добавляем новое сообщение в поток
                            TelegramMessage.objects.create(
                                message_id=str(result.message_id),
                                chat_id=str(ticket.telegram_chat_id),
                                chat_title=ticket.telegram_chat_title or '',
                                from_user_id=str(request.user.id),
                                from_username=request.user.username,
                                from_fullname=request.user.get_full_name() or request.user.username,
                                text=new_content,
                                message_date=timezone.now(),
                                created_at=timezone.now(),
                                linked_ticket=ticket,
                                linked_action='add_comment',
                                reply_to_message_id=str(ticket.external_message_id)
                            )
                            logger.info(f"Added new comment message to stream: {result.message_id}")
                    except Exception as stream_error:
                        logger.error(f"Failed to update comment message in stream: {stream_error}", exc_info=True)
                    
                    messages.success(request, 'Комментарий обновлен и отправлен в Telegram')
                else:
                    logger.error("TELEGRAM_BOT_TOKEN not configured")
                    messages.warning(request, 'Telegram бот не настроен')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to edit Telegram comment: {e}", exc_info=True)
                messages.warning(request, f'Не удалось обновить сообщение в Telegram: {str(e)}')
        else:
            messages.success(request, 'Комментарий обновлен')
        
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Показываем форму редактирования
    context = {
        'comment': comment,
        'ticket': ticket,
        'can_reply_in_telegram': can_reply_in_telegram,
    }
    
    return render(request, 'tickets/ticket_edit_comment.html', context)


@login_required
def delete_comment(request, comment_id):
    """Удалить комментарий"""
    comment = get_object_or_404(TicketComment, id=comment_id)
    ticket = comment.ticket
    
    # Проверяем, что комментарий был отправлен в Telegram
    if not comment.telegram_message_id:
        messages.error(request, 'Удаление недоступно для данного комментария')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что обращение имеет Telegram данные
    if not ticket.telegram_chat_id or not ticket.external_message_id:
        messages.error(request, 'Удаление недоступно для данного обращения')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что группа разрешена для отправки ответов
    can_reply_in_telegram = False
    try:
        telegram_group = TelegramGroup.objects.get(chat_id=ticket.telegram_chat_id)
        can_reply_in_telegram = not telegram_group.is_blocked
    except TelegramGroup.DoesNotExist:
        can_reply_in_telegram = False
    
    if not can_reply_in_telegram:
        messages.error(request, 'Удаление недоступно для данной группы')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что комментарий отправлен ботом (от пользователя системы)
    if comment.author_type != 'user':
        messages.error(request, 'Удаление недоступно для комментариев клиентов')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    if request.method == 'POST':
        try:
            import logging
            import asyncio
            from telegram.ext import Application
            from django.conf import settings
            
            logger = logging.getLogger(__name__)
            logger.info(f"Attempting to delete Telegram comment: chat_id={ticket.telegram_chat_id}, message_id={comment.telegram_message_id}")
            
            bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
            if bot_token:
                # Создаем асинхронную функцию для удаления сообщения
                async def delete_telegram_message():
                    application = Application.builder().token(bot_token).build()
                    result = await application.bot.delete_message(
                        chat_id=ticket.telegram_chat_id,
                        message_id=int(comment.telegram_message_id)
                    )
                    return result
                
                # Запускаем асинхронную функцию
                result = asyncio.run(delete_telegram_message())
                logger.info(f"Telegram message deleted successfully: {comment.telegram_message_id}")
                
                # Удаляем сообщение из потока
                try:
                    telegram_message = TelegramMessage.objects.filter(
                        message_id=comment.telegram_message_id,
                        chat_id=ticket.telegram_chat_id
                    ).delete()
                    logger.info(f"Deleted message from stream: {comment.telegram_message_id}")
                except Exception as stream_error:
                    logger.error(f"Failed to delete message from stream: {stream_error}", exc_info=True)
                
                # Создаем запись аудита
                TicketAudit.objects.create(
                    ticket=ticket,
                    action='comment_deleted',
                    user=request.user,
                    comment=f'Комментарий удален: {comment.content[:50]}...'
                )
                
                # Удаляем комментарий
                comment.delete()
                
                messages.success(request, 'Комментарий удален из Telegram и потока')
            else:
                logger.error("TELEGRAM_BOT_TOKEN not configured")
                
                # Удаляем только комментарий и сообщение из потока
                TelegramMessage.objects.filter(
                    message_id=comment.telegram_message_id,
                    chat_id=ticket.telegram_chat_id
                ).delete()
                comment.delete()
                
                messages.warning(request, 'Комментарий удален (Telegram бот не настроен)')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to delete Telegram comment: {e}", exc_info=True)
            messages.warning(request, f'Не удалось удалить сообщение из Telegram: {str(e)}')
        
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Показываем форму подтверждения удаления
    context = {
        'comment': comment,
        'ticket': ticket,
    }
    
    return render(request, 'tickets/ticket_delete_comment.html', context)


@login_required
def delete_resolution(request, ticket_id):
    """Удалить решение обращения"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Проверяем, что обращение решено и имеет Telegram данные
    if not ticket.status.is_final or not ticket.telegram_chat_id or not ticket.external_message_id:
        messages.error(request, 'Удаление решения недоступно для данного обращения')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что группа разрешена для отправки ответов
    can_reply_in_telegram = False
    try:
        telegram_group = TelegramGroup.objects.get(chat_id=ticket.telegram_chat_id)
        can_reply_in_telegram = not telegram_group.is_blocked
    except TelegramGroup.DoesNotExist:
        can_reply_in_telegram = False
    
    if not can_reply_in_telegram:
        messages.error(request, 'Удаление решения недоступно для данной группы')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    if request.method == 'POST':
        try:
            import logging
            import asyncio
            from telegram.ext import Application
            from django.conf import settings
            
            logger = logging.getLogger(__name__)
            # Находим сообщение с решением в потоке
            resolution_message = TelegramMessage.objects.filter(
                chat_id=ticket.telegram_chat_id,
                linked_ticket=ticket,
                linked_action='resolve_ticket'
            ).first()
            
            if not resolution_message:
                messages.error(request, 'Сообщение с решением не найдено в потоке')
                return redirect('tickets:ticket_detail', ticket_id=ticket.id)
            
            logger.info(f"Attempting to delete Telegram resolution: chat_id={ticket.telegram_chat_id}, message_id={resolution_message.message_id}")
            
            bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
            if bot_token:
                # Создаем асинхронную функцию для удаления сообщения
                async def delete_telegram_message():
                    application = Application.builder().token(bot_token).build()
                    result = await application.bot.delete_message(
                        chat_id=ticket.telegram_chat_id,
                        message_id=int(resolution_message.message_id)
                    )
                    return result
                
                # Запускаем асинхронную функцию
                result = asyncio.run(delete_telegram_message())
                logger.info(f"Telegram message deleted successfully: {resolution_message.message_id}")
                
                # Удаляем сообщение из потока
                try:
                    resolution_message.delete()
                    logger.info(f"Deleted resolution message from stream: {resolution_message.message_id}")
                except Exception as stream_error:
                    logger.error(f"Failed to delete resolution message from stream: {stream_error}", exc_info=True)
                
                # Создаем запись аудита
                TicketAudit.objects.create(
                    ticket=ticket,
                    action='resolution_deleted',
                    user=request.user,
                    comment=f'Решение удалено: {ticket.resolution[:50]}...'
                )
                
                # Очищаем решение и меняем статус
                ticket.resolution = ''
                ticket.resolved_at = None
                ticket.status = TicketStatus.objects.get(name='В работе')  # Возвращаем в работу
                ticket.save()
                
                messages.success(request, 'Решение удалено из Telegram и потока')
            else:
                logger.error("TELEGRAM_BOT_TOKEN not configured")
                
                # Удаляем только сообщение из потока и очищаем решение
                if resolution_message:
                    resolution_message.delete()
                
                ticket.resolution = ''
                ticket.resolved_at = None
                ticket.status = TicketStatus.objects.get(name='В работе')
                ticket.save()
                
                messages.warning(request, 'Решение удалено (Telegram бот не настроен)')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to delete Telegram resolution: {e}", exc_info=True)
            messages.warning(request, f'Не удалось удалить сообщение из Telegram: {str(e)}')
        
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Показываем форму подтверждения удаления
    context = {
        'ticket': ticket,
        'can_reply_in_telegram': can_reply_in_telegram,
    }
    
    return render(request, 'tickets/ticket_delete_resolution.html', context)


@login_required
def ticket_create(request):
    """Создание нового обращения"""
    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            
            # Обрабатываем autocomplete поля
            category_id = request.POST.get('category_id')
            client_id = request.POST.get('client_id')
            organization_id = request.POST.get('organization_id')
            assigned_to_text = (request.POST.get('assigned_to') or '').strip()
            assigned_to_id_raw = (request.POST.get('assigned_to_id') or '').strip()
            
            if category_id:
                try:
                    ticket.category = Category.objects.get(id=category_id)
                except Category.DoesNotExist:
                    pass
            
            if client_id:
                try:
                    ticket.client = Client.objects.get(id=client_id)
                except Client.DoesNotExist:
                    pass
            
            if organization_id:
                try:
                    ticket.organization = Organization.objects.get(id=organization_id)
                except Organization.DoesNotExist:
                    pass
            
            if not assigned_to_text:
                ticket.assigned_to = None
            elif assigned_to_id_raw.isdigit():
                try:
                    ticket.assigned_to = User.objects.get(id=int(assigned_to_id_raw))
                except User.DoesNotExist:
                    ticket.assigned_to = None
            else:
                # Если hidden assigned_to_id отсутствует (пользователь очистил поле) — снимаем исполнителя
                ticket.assigned_to = None
            
            ticket.save()
            
            # Обрабатываем вложения
            files = request.FILES.getlist('attachments')
            uploaded_count = 0
            
            for file in files:
                attachment = TicketAttachment(
                    ticket=ticket,
                    file=file,
                    filename=file.name,
                    file_size=file.size,
                    uploaded_by=request.user
                )
                attachment.save()
                uploaded_count += 1
            
            # Создаем запись аудита
            audit_comment = 'Обращение создано'
            if uploaded_count > 0:
                audit_comment += f' с {uploaded_count} вложениями'
            
            TicketAudit.objects.create(
                ticket=ticket,
                action='created',
                user=request.user,
                comment=audit_comment
            )
            
            messages.success(request, f'Обращение #{ticket.id} создано')
            return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    else:
        form = TicketForm()
        # Предзаполнение исполнителя текущим пользователем в UI (без фиксации на сервере)
        # Текстовое значение — имя/логин, скрытый id будет создан JS при первом выборе из списка,
        # поэтому сервер учтёт очистку/замену корректно.
        form.fields['assigned_to'].initial = request.user.get_full_name() or request.user.username
    
    context = {
        'form': form,
        'categories_for_sla': Category.objects.filter(is_active=True).order_by('sla_hours', 'name')[:6],
    }
    
    return render(request, 'tickets/ticket_form.html', context)


@login_required
def ticket_edit(request, ticket_id):
    """Редактирование обращения"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if request.method == 'POST':
        form = TicketForm(request.POST, instance=ticket)
        if form.is_valid():
            old_status = ticket.status
            old_assigned = ticket.assigned_to
            
            ticket = form.save(commit=False)
            
            # Обрабатываем autocomplete поля
            category_id = request.POST.get('category_id')
            client_id = request.POST.get('client_id')
            organization_id = request.POST.get('organization_id')
            assigned_to_text = (request.POST.get('assigned_to') or '').strip()
            assigned_to_id_raw = (request.POST.get('assigned_to_id') or '').strip()
            
            if category_id:
                try:
                    ticket.category = Category.objects.get(id=category_id)
                except Category.DoesNotExist:
                    pass
            
            if client_id:
                try:
                    ticket.client = Client.objects.get(id=client_id)
                except Client.DoesNotExist:
                    pass
            
            if organization_id:
                try:
                    ticket.organization = Organization.objects.get(id=organization_id)
                except Organization.DoesNotExist:
                    pass
            
            if not assigned_to_text:
                ticket.assigned_to = None
            elif assigned_to_id_raw.isdigit():
                try:
                    ticket.assigned_to = User.objects.get(id=int(assigned_to_id_raw))
                except User.DoesNotExist:
                    ticket.assigned_to = None
            else:
                ticket.assigned_to = None
            
            ticket.save()
            
            # Создаем записи аудита для изменений
            if old_status != ticket.status:
                TicketAudit.objects.create(
                    ticket=ticket,
                    action='status_changed',
                    user=request.user,
                    old_value=old_status.name,
                    new_value=ticket.status.name,
                    comment='Статус изменен'
                )
            
            if old_assigned != ticket.assigned_to:
                TicketAudit.objects.create(
                    ticket=ticket,
                    action='assigned',
                    user=request.user,
                    old_value=old_assigned.username if old_assigned else 'Не назначен',
                    new_value=ticket.assigned_to.username if ticket.assigned_to else 'Не назначен',
                    comment='Исполнитель изменен'
                )
            
            messages.success(request, f'Обращение #{ticket.id} обновлено')
            return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    else:
        form = TicketForm(instance=ticket)
        # Предзаполняем текстовые поля и скрытые id значения
        if ticket.category:
            form.fields['category'].initial = str(ticket.category)
        if ticket.client:
            form.fields['client'].initial = ticket.client.name
        if ticket.assigned_to:
            form.fields['assigned_to'].initial = ticket.assigned_to.get_full_name() or ticket.assigned_to.username
    
    context = {
        'form': form,
        'ticket': ticket,
        'categories_for_sla': Category.objects.filter(is_active=True).order_by('sla_hours', 'name')[:6],
    }
    
    return render(request, 'tickets/ticket_form.html', context)


@login_required
def take_ticket(request, ticket_id):
    """Взять обращение в работу или вернуть из финального статуса"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Проверяем, можно ли взять в работу
    if ticket.assigned_to and ticket.assigned_to != request.user and not ticket.status.is_final:
        messages.error(request, 'Обращение уже назначено другому исполнителю')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Находим статус "В работе"
    working_status = TicketStatus.objects.filter(is_working=True).first()
    if not working_status:
        messages.error(request, 'Статус "В работе" не найден')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Определяем действие для аудита
    if ticket.status.is_final:
        action = 'returned_to_work'
        comment_text = 'Возвращено в работу из финального статуса'
        success_message = f'Обращение #{ticket.id} возвращено в работу'
    else:
        action = 'taken'
        comment_text = 'Взято в работу'
        success_message = f'Обращение #{ticket.id} взято в работу'
    
    ticket.assigned_to = request.user
    ticket.status = working_status
    ticket.taken_at = timezone.now()
    ticket.save()
    
    # Создаем запись аудита
    TicketAudit.objects.create(
        ticket=ticket,
        action=action,
        user=request.user,
        comment=comment_text
    )
    
    messages.success(request, success_message)
    return redirect('tickets:ticket_detail', ticket_id=ticket.id)


@login_required
def resolve_ticket(request, ticket_id):
    """Решить обращение"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if request.method == 'POST':
        resolution = request.POST.get('resolution', '')
        resolution_notes = request.POST.get('resolution_notes', '')
        reply_in_chat = request.POST.get('reply_in_chat') == '1'
        
        # Находим статус "Решено"
        resolved_status = TicketStatus.objects.filter(name='Решено').first()
        if not resolved_status:
            messages.error(request, 'Статус "Решено" не найден')
            return redirect('tickets:ticket_detail', ticket_id=ticket.id)
        
        ticket.status = resolved_status
        ticket.resolution = resolution
        ticket.resolution_notes = resolution_notes
        ticket.resolved_at = timezone.now()
        
        # Если taken_at пустое, устанавливаем на 1 секунду раньше resolved_at
        if not ticket.taken_at:
            ticket.taken_at = ticket.resolved_at - timezone.timedelta(seconds=1)
        
        ticket.save()
        
        # Создаем запись аудита
        audit_comment = f'Решено: {resolution[:50]}...' if resolution else 'Решено'
        if reply_in_chat and ticket.telegram_chat_id and ticket.external_message_id:
            audit_comment += ' (ответ отправлен в Telegram)'
        
        TicketAudit.objects.create(
            ticket=ticket,
            action='resolved',
            user=request.user,
            comment=audit_comment
        )
        
        # Отправляем ответ в Telegram, если запрошено
        if reply_in_chat and ticket.telegram_chat_id and ticket.external_message_id:
            try:
                import logging
                import asyncio
                from telegram.ext import Application
                from django.conf import settings
                
                logger = logging.getLogger(__name__)
                logger.info(f"Attempting to send Telegram reply: chat_id={ticket.telegram_chat_id}, message_id={ticket.external_message_id}")
                
                bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                if bot_token:
                    # Создаем асинхронную функцию для отправки
                    async def send_telegram_message():
                        application = Application.builder().token(bot_token).build()
                        result = await application.bot.send_message(
                            chat_id=ticket.telegram_chat_id,
                            text=resolution,
                            reply_to_message_id=int(ticket.external_message_id)
                        )
                        return result
                    
                    # Запускаем асинхронную функцию
                    result = asyncio.run(send_telegram_message())
                    logger.info(f"Telegram message sent successfully: {result.message_id}")
                    
                    # Добавляем отправленное сообщение в поток
                    try:
                        from tickets.models import TelegramMessage
                        from django.contrib.auth.models import User
                        
                        # Создаем запись в потоке о том, что обращение решено
                        TelegramMessage.objects.create(
                            message_id=str(result.message_id),
                            chat_id=str(ticket.telegram_chat_id),
                            chat_title=ticket.telegram_chat_title or '',
                            from_user_id=str(request.user.id),
                            from_username=request.user.username,
                            from_fullname=request.user.get_full_name() or request.user.username,
                            text=resolution,
                            message_date=timezone.now(),
                            created_at=timezone.now(),
                            linked_ticket=ticket,
                            linked_action='resolve_ticket',
                            reply_to_message_id=str(ticket.external_message_id)
                        )
                        logger.info(f"Added resolution message to stream: {result.message_id}")
                    except Exception as stream_error:
                        logger.error(f"Failed to add resolution message to stream: {stream_error}", exc_info=True)
                    
                    messages.success(request, 'Ответ отправлен в Telegram')
                else:
                    logger.error("TELEGRAM_BOT_TOKEN not configured")
                    messages.warning(request, 'Telegram бот не настроен')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to send Telegram reply: {e}", exc_info=True)
                messages.warning(request, f'Не удалось отправить ответ в Telegram: {str(e)}')
        
        messages.success(request, f'Обращение #{ticket.id} решено')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Показываем форму решения
    templates = TicketTemplate.objects.filter(
        category=ticket.category,
        is_active=True
    )
    
    # Проверяем, можно ли отправлять ответ в Telegram
    can_reply_in_telegram = False
    if ticket.telegram_chat_id and ticket.external_message_id:
        # Проверяем, есть ли группа в нашей модели и не заблокирована ли она
        try:
            telegram_group = TelegramGroup.objects.get(chat_id=ticket.telegram_chat_id)
            can_reply_in_telegram = not telegram_group.is_blocked
        except TelegramGroup.DoesNotExist:
            # Если группы нет в нашей модели, не показываем галку
            can_reply_in_telegram = False
    
    context = {
        'ticket': ticket,
        'templates': templates,
        'can_reply_in_telegram': can_reply_in_telegram,
    }
    
    return render(request, 'tickets/ticket_resolve.html', context)


@login_required
def edit_resolution(request, ticket_id):
    """Редактировать решение обращения"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Проверяем, что обращение решено и имеет Telegram данные
    if not ticket.status.is_final or not ticket.telegram_chat_id or not ticket.external_message_id:
        messages.error(request, 'Редактирование решения недоступно для данного обращения')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Проверяем, что группа разрешена для отправки ответов
    can_reply_in_telegram = False
    try:
        telegram_group = TelegramGroup.objects.get(chat_id=ticket.telegram_chat_id)
        can_reply_in_telegram = not telegram_group.is_blocked
    except TelegramGroup.DoesNotExist:
        can_reply_in_telegram = False
    
    if not can_reply_in_telegram:
        messages.error(request, 'Редактирование решения недоступно для данной группы')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    if request.method == 'POST':
        new_resolution = request.POST.get('resolution', '')
        reply_in_chat = request.POST.get('reply_in_chat') == '1'
        
        if not new_resolution:
            messages.error(request, 'Решение не может быть пустым')
            return redirect('tickets:edit_resolution', ticket_id=ticket.id)
        
        # Обновляем решение
        ticket.resolution = new_resolution
        ticket.save()
        
        # Создаем запись аудита
        TicketAudit.objects.create(
            ticket=ticket,
            action='resolution_edited',
            user=request.user,
            comment=f'Решение отредактировано: {new_resolution[:50]}...'
        )
        
        # Если запрошена отправка в Telegram, обновляем сообщение
        if reply_in_chat:
            try:
                import logging
                import asyncio
                from telegram.ext import Application
                from django.conf import settings
                
                logger = logging.getLogger(__name__)
                # Находим сообщение с решением в потоке
                resolution_message = TelegramMessage.objects.filter(
                    chat_id=ticket.telegram_chat_id,
                    linked_ticket=ticket,
                    linked_action='resolve_ticket'
                ).first()
                
                if not resolution_message:
                    messages.error(request, 'Сообщение с решением не найдено в потоке')
                    return redirect('tickets:ticket_detail', ticket_id=ticket.id)
                
                logger.info(f"Attempting to edit Telegram message: chat_id={ticket.telegram_chat_id}, message_id={resolution_message.message_id}")
                
                bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                if bot_token:
                    # Создаем асинхронную функцию для редактирования сообщения
                    async def edit_telegram_message():
                        application = Application.builder().token(bot_token).build()
                        try:
                            # Пытаемся отредактировать существующее сообщение с решением
                            result = await application.bot.edit_message_text(
                                chat_id=ticket.telegram_chat_id,
                                message_id=int(resolution_message.message_id),
                                text=new_resolution
                            )
                            return result, 'edited'
                        except Exception as edit_error:
                            # Если не удалось отредактировать, удаляем старое сообщение и отправляем новое
                            logger.warning(f"Could not edit message, deleting old and sending new one: {edit_error}")
                            try:
                                await application.bot.delete_message(
                                    chat_id=ticket.telegram_chat_id,
                                    message_id=int(resolution_message.message_id)
                                )
                                logger.info(f"Deleted old resolution message: {resolution_message.message_id}")
                            except Exception as delete_error:
                                logger.warning(f"Could not delete old message: {delete_error}")
                            
                            result = await application.bot.send_message(
                                chat_id=ticket.telegram_chat_id,
                                text=new_resolution,
                                reply_to_message_id=int(ticket.external_message_id)
                            )
                            return result, 'new'
                    
                    # Запускаем асинхронную функцию
                    result, action_type = asyncio.run(edit_telegram_message())
                    logger.info(f"Telegram message {action_type} successfully: {result.message_id}")
                    
                    # Обновляем сообщение в потоке
                    try:
                        if action_type == 'edited':
                            # Обновляем существующее сообщение
                            telegram_message = TelegramMessage.objects.filter(
                                chat_id=ticket.telegram_chat_id,
                                linked_ticket=ticket,
                                linked_action='resolve_ticket'
                            ).first()
                            
                            if telegram_message:
                                telegram_message.text = new_resolution
                                telegram_message.save()
                                logger.info(f"Updated resolution message in stream: {telegram_message.id}")
                        else:
                            # Удаляем старое сообщение из потока и добавляем новое
                            old_message_id = resolution_message.message_id
                            resolution_message.delete()
                            logger.info(f"Deleted old resolution message from stream: {old_message_id}")
                            
                            TelegramMessage.objects.create(
                                message_id=str(result.message_id),
                                chat_id=str(ticket.telegram_chat_id),
                                chat_title=ticket.telegram_chat_title or '',
                                from_user_id=str(request.user.id),
                                from_username=request.user.username,
                                from_fullname=request.user.get_full_name() or request.user.username,
                                text=new_resolution,
                                message_date=timezone.now(),
                                created_at=timezone.now(),
                                linked_ticket=ticket,
                                linked_action='resolve_ticket',
                                reply_to_message_id=str(ticket.external_message_id)
                            )
                            logger.info(f"Added new resolution message to stream: {result.message_id}")
                    except Exception as stream_error:
                        logger.error(f"Failed to update resolution message in stream: {stream_error}", exc_info=True)
                    
                    messages.success(request, 'Решение обновлено и отправлено в Telegram')
                else:
                    logger.error("TELEGRAM_BOT_TOKEN not configured")
                    messages.warning(request, 'Telegram бот не настроен')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to edit Telegram message: {e}", exc_info=True)
                messages.warning(request, f'Не удалось обновить сообщение в Telegram: {str(e)}')
        else:
            messages.success(request, 'Решение обновлено')
        
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Показываем форму редактирования
    context = {
        'ticket': ticket,
        'can_reply_in_telegram': can_reply_in_telegram,
    }
    
    return render(request, 'tickets/ticket_edit_resolution.html', context)


@login_required
def close_ticket(request, ticket_id):
    """Закрыть обращение (подтверждено заявителем) → считаем как Решено"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Находим статус "Решено" (единственный финальный)
    resolved_status = TicketStatus.objects.filter(name='Решено').first()
    if not resolved_status:
        resolved_status = TicketStatus.objects.filter(is_final=True).first()
    if not resolved_status:
        messages.error(request, 'Финальный статус "Решено" не найден')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    
    # Обновляем обращение
    ticket.status = resolved_status
    if not ticket.resolved_at:
        ticket.resolved_at = timezone.now()
    
    # Если taken_at пустое, устанавливаем на 1 секунду раньше resolved_at
    if not ticket.taken_at:
        ticket.taken_at = ticket.resolved_at - timezone.timedelta(seconds=1)
    
    ticket.save()
    
    # Аудит
    TicketAudit.objects.create(
        ticket=ticket,
        action='resolved',
        user=request.user,
        comment='Обращение подтверждено заявителем (Решено)'
    )
    
    messages.success(request, f'Обращение #{ticket.id} решено')
    return redirect('tickets:ticket_detail', ticket_id=ticket.id)


@login_required
def set_waiting(request, ticket_id):
    """Перевести статус в "Ожидает ответа". Разрешено только из статусов с is_working=True и если уже не в этом статусе."""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    if not ticket.status.is_working:
        messages.error(request, 'Перевод возможен только из статуса "В работе"')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    waiting_status = TicketStatus.objects.filter(name='Ожидает ответа').first()
    if not waiting_status:
        messages.error(request, 'Статус "Ожидает ответа" не найден')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    if ticket.status == waiting_status:
        messages.info(request, 'Статус уже "Ожидает ответа"')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    old_status = ticket.status
    ticket.status = waiting_status
    ticket.save()
    TicketAudit.objects.create(
        ticket=ticket,
        action='status_changed',
        user=request.user,
        old_value=old_status.name,
        new_value=waiting_status.name,
        comment='Переведено в статус Ожидает ответа'
    )
    messages.success(request, f'Обращение #{ticket.id} переведено в статус "Ожидает ответа"')
    return redirect('tickets:ticket_detail', ticket_id=ticket.id)


@login_required
def return_to_work(request, ticket_id):
    """Вернуть из "Ожидает ответа" в рабочий статус."""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    if ticket.status and ticket.status.name != 'Ожидает ответа':
        messages.error(request, 'Возврат возможен только из статуса "Ожидает ответа"')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    working_status = TicketStatus.objects.filter(is_working=True).exclude(name='Ожидает ответа').first()
    if not working_status:
        messages.error(request, 'Рабочий статус не найден')
        return redirect('tickets:ticket_detail', ticket_id=ticket.id)
    old_status = ticket.status
    ticket.status = working_status
    # taken_at оставляем как есть; исполнитель не меняется
    ticket.save()
    TicketAudit.objects.create(
        ticket=ticket,
        action='status_changed',
        user=request.user,
        old_value=old_status.name if old_status else '',
        new_value=working_status.name,
        comment='Возвращено в работу из статуса Ожидает ответа'
    )
    messages.success(request, f'Обращение #{ticket.id} возвращено в работу')
    return redirect('tickets:ticket_detail', ticket_id=ticket.id)

@login_required
def client_list(request):
    """Список клиентов"""
    search_query = request.GET.get('search', '')
    clients = Client.objects.filter(is_active=True).annotate(
        ticket_count=Count('ticket')
    ).order_by('name')
    
    if search_query:
        clients = clients.filter(
            Q(name__iregex=search_query) |
            Q(contact_person__iregex=search_query)
        )
    
    # Пагинация
    paginator = Paginator(clients, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_count': clients.count(),
    }
    
    return render(request, 'tickets/client_list.html', context)


@login_required
def client_detail(request, client_id):
    """Детальная страница клиента"""
    client = get_object_or_404(Client, id=client_id)
    tickets = client.ticket_set.select_related(
        'category', 'status', 'assigned_to'
    ).order_by('-created_at')
    # Статистика
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status__is_final=False).count()
    
    context = {
        'client': client,
        'tickets': tickets,
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
    }
    
    return render(request, 'tickets/client_detail.html', context)


@login_required
def client_create(request):
    """Создание нового клиента"""
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save()
            messages.success(request, f'Клиент "{client.name}" создан')
            return redirect('tickets:client_detail', client_id=client.id)
    else:
        form = ClientForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'tickets/client_form.html', context)


@login_required
def client_edit(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            client = form.save()
            messages.success(request, f'Клиент "{client.name}" обновлён')
            return redirect('tickets:client_detail', client_id=client.id)
    else:
        form = ClientForm(instance=client)
    return render(request, 'tickets/client_form.html', {'form': form, 'client': client})


@login_required
def get_template_content(request, template_id):
    """AJAX запрос для получения содержимого шаблона"""
    template = get_object_or_404(TicketTemplate, id=template_id)
    
    return JsonResponse({
        'title': template.title_template,
        'content': template.content_template
    })


@login_required
def autocomplete_categories(request):
    """AJAX autocomplete для категорий"""
    query = request.GET.get('q', '')
    base_qs = Category.objects.filter(is_active=True)
    if len(query) >= 1:
        base_qs = base_qs.filter(Q(name__icontains=query) | Q(description__icontains=query))
    categories = base_qs.order_by('parent__name', 'name')[:10]
    
    results = []
    for category in categories:
        results.append({
            'id': category.id,
            'text': str(category),
            'parent': category.parent.name if category.parent else None
        })
    
    return JsonResponse({'results': results})


@login_required
def autocomplete_clients(request):
    """AJAX autocomplete для клиентов"""
    query = request.GET.get('q', '')
    base_qs = Client.objects.filter(is_active=True)
    if len(query) >= 1:
        # Используем iregex для регистронезависимого поиска в SQLite
        base_qs = base_qs.filter(
            Q(name__iregex=query) |
            Q(contact_person__iregex=query) |
            Q(phone__iregex=query) |
            Q(email__iregex=query) |
            Q(organization__name__iregex=query)
        )
    clients = base_qs.select_related('organization').order_by('name')[:10]
    
    results = []
    for client in clients:
        display = client.name
        if client.organization:
            display = f"{client.organization.name} — {client.name}"
        results.append({
            'id': client.id,
            'text': display,
            'contact_person': client.contact_person,
            'phone': client.phone,
            'email': client.email
        })
    
    return JsonResponse({'results': results})


@login_required
def autocomplete_organizations(request):
    """AJAX autocomplete для организаций"""
    query = request.GET.get('q', '')
    base_qs = Organization.objects.all()
    if len(query) >= 1:
        # Используем iregex для регистронезависимого поиска в SQLite
        base_qs = base_qs.filter(name__iregex=query)
    orgs = base_qs.order_by('name')[:10]
    results = [{'id': o.id, 'text': o.name} for o in orgs]
    return JsonResponse({'results': results})


@login_required
def create_organization(request):
    """Создание организации (AJAX). Принимает name, возвращает {id, name}."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'error': 'Введите название организации'}, status=400)
    org, created = Organization.objects.get_or_create(name=name)
    return JsonResponse({'id': org.id, 'name': org.name, 'created': created})


@login_required
def autocomplete_users(request):
    """AJAX autocomplete для пользователей"""
    query = request.GET.get('q', '')
    base_qs = User.objects.filter(is_active=True)
    if len(query) >= 1:
        base_qs = base_qs.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    users = base_qs.order_by('username')[:10]
    
    results = []
    for user in users:
        full_name = user.get_full_name()
        display_name = full_name if full_name else user.username
        results.append({
            'id': user.id,
            'text': display_name,
            'username': user.username,
            'email': user.email
        })
    
    return JsonResponse({'results': results})


def autocomplete_groups(request):
    """AJAX autocomplete для групп Telegram"""
    from .models import TelegramGroup
    
    query = request.GET.get('q', '')
    base_qs = TelegramGroup.objects.all()
    if len(query) >= 1:
        base_qs = base_qs.filter(
            Q(title__iregex=query) |
            Q(chat_id__iregex=query)
        )
    groups = base_qs.order_by('title', 'chat_id')[:10]
    
    results = []
    for group in groups:
        results.append({
            'id': group.chat_id,
            'text': group.title or group.chat_id,
            'chat_id': group.chat_id,
            'title': group.title
        })
    
    return JsonResponse({'results': results})


@login_required
def delete_attachment(request, attachment_id):
    """Удаление вложения"""
    attachment = get_object_or_404(TicketAttachment, id=attachment_id)
    ticket_id = attachment.ticket.id
    
    # Проверяем права доступа
    if attachment.uploaded_by != request.user and not request.user.is_staff:
        messages.error(request, 'У вас нет прав для удаления этого файла')
        return redirect('tickets:ticket_detail', ticket_id=ticket_id)
    
    # Создаем запись аудита
    TicketAudit.objects.create(
        ticket=attachment.ticket,
        action='updated',
        user=request.user,
        comment=f'Удален файл: {attachment.filename}'
    )
    
    attachment.delete()
    messages.success(request, f'Файл "{attachment.filename}" удален')
    return redirect('tickets:ticket_detail', ticket_id=ticket_id)


@login_required
def queue_view(request):
    """Очередь дел"""
    # Не назначенные обращения
    unassigned_tickets = Ticket.objects.filter(
        assigned_to__isnull=True,
        status__is_final=False
    ).select_related('client', 'category', 'status').order_by('created_at')
    
    # Мои обращения
    my_tickets = Ticket.objects.filter(
        assigned_to=request.user,
        status__is_final=False
    ).select_related('client', 'category', 'status').order_by('created_at')
    
    # Просроченные обращения
    overdue_tickets = Ticket.objects.filter(
        status__is_final=False,
        created_at__lt=timezone.now() - timezone.timedelta(hours=24)
    ).select_related('client', 'category', 'status', 'assigned_to').order_by('created_at')
    
    context = {
        'unassigned_tickets': unassigned_tickets,
        'my_tickets': my_tickets,
        'overdue_tickets': overdue_tickets,
    }
    
    return render(request, 'tickets/queue.html', context)


@login_required
def analytics(request):
    """Страница аналитики обращений"""
    # Утилита для извлечения первого ненулевого значения параметра (учитывая дубли)
    def first_non_empty(param_name):
        values = request.GET.getlist(param_name)
        for v in values:
            if v not in (None, '', 'None', 'null', 'NULL'):
                return v
        return None

    # Фильтры
    category_id = first_non_empty('category_id')
    category_text = first_non_empty('category')
    client_id = first_non_empty('client_id') or first_non_empty('client')
    organization_id = first_non_empty('organization_id') or first_non_empty('organization')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    chart_type = request.GET.get('chart_type', 'organizations')  # tags или organizations

    # Даты по умолчанию: текущий месяц
    from django.utils import timezone as dj_tz
    today = dj_tz.localdate()
    if not date_from:
        first_day = today.replace(day=1)
        date_from = first_day.isoformat()
    if not date_to:
        date_to = today.isoformat()

    tickets_qs = Ticket.objects.select_related('category', 'client', 'organization', 'status', 'assigned_to').all()

    if category_id and str(category_id).isdigit():
        cid = int(category_id)
        tickets_qs = tickets_qs.filter(Q(category_id=cid) | Q(category__parent_id=cid))
    elif category_text:
        tickets_qs = tickets_qs.filter(
            Q(category__name__icontains=category_text) |
            Q(category__parent__name__icontains=category_text)
        )
    if client_id and str(client_id).isdigit():
        tickets_qs = tickets_qs.filter(client_id=int(client_id))
    if organization_id and str(organization_id).isdigit():
        tickets_qs = tickets_qs.filter(organization_id=int(organization_id))
    if date_from:
        tickets_qs = tickets_qs.filter(created_at__date__gte=date_from)
    if date_to:
        tickets_qs = tickets_qs.filter(created_at__date__lte=date_to)

    # Агрегации (без extra, чтобы избежать конфликтов алиасов)
    by_day = (
        tickets_qs
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .order_by('day')
        .annotate(cnt=Count('id'))
    )
    # Подсчет тегов или организаций в зависимости от chart_type
    if chart_type == 'organizations':
        from collections import Counter
        org_counter = Counter()
        for ticket in tickets_qs.select_related('organization'):
            if ticket.organization:
                org_counter[ticket.organization.name] += 1
        top_orgs = sorted(org_counter.items(), key=lambda x: x[1], reverse=True)[:20]
        chart_data = [{'name': k, 'count': v} for k, v in top_orgs]
    else:  # tags
        by_tags = []
        # Разбор тегов по запятым
        for t in tickets_qs.exclude(tags="").values_list('tags', flat=True):
            for tag in [x.strip() for x in t.split(',') if x.strip()]:
                by_tags.append(tag)
        from collections import Counter
        tags_counter = Counter(by_tags)
        top_tags = sorted(tags_counter.items(), key=lambda x: x[1], reverse=True)[:20]
        chart_data = [{'name': k, 'count': v} for k, v in top_tags]

    # Сериализация для фронта
    by_day_list = list(by_day)
    by_day_serialized = [
        {'day': (item['day'].isoformat() if item['day'] else None), 'cnt': item['cnt']}
        for item in by_day_list
    ]
    import json
    chart_by_day_json = json.dumps(by_day_serialized)
    chart_data_json = json.dumps(chart_data)

    # Краткая сводка для подписи под графиком
    total_count = tickets_qs.count()
    day_count = len(by_day_list)
    avg_per_day = round(total_count / day_count, 1) if day_count else 0

    # Пагинация списка
    paginator = Paginator(tickets_qs.order_by('-created_at'), 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Тексты выбранных значений для автозаполнения
    category_name = None
    client_name = None
    organization_name = None
    if category_id and str(category_id).isdigit():
        c = Category.objects.filter(id=int(category_id)).first()
        if c:
            category_name = str(c)
    if client_id and str(client_id).isdigit():
        cl = Client.objects.filter(id=int(client_id)).first()
        if cl:
            client_name = cl.name
    if organization_id and str(organization_id).isdigit():
        org = Organization.objects.filter(id=int(organization_id)).first()
        if org:
            organization_name = org.name

    context = {
        'page_obj': page_obj,
        'filters': {
            'category_id': category_id or '',
            'category_name': category_name or category_text or '',
            'client_id': client_id or '',
            'client_name': client_name or '',
            'organization_id': organization_id or '',
            'organization_name': organization_name or '',
            'date_from': date_from or '',
            'date_to': date_to or '',
        },
        'chart_by_day_json': chart_by_day_json,
        'chart_data_json': chart_data_json,
        'chart_type': chart_type,
        'analytics_summary': {
            'total_count': total_count,
            'day_count': day_count,
            'avg_per_day': avg_per_day,
        },
    }

    return render(request, 'tickets/analytics.html', context)


@login_required
def analytics_export_xlsx(request):
    """Экспорт выборки аналитики в XLSX по текущим фильтрам."""
    # Повторяем фильтрацию как в analytics
    def first_non_empty(param_name):
        values = request.GET.getlist(param_name)
        for v in values:
            if v not in (None, '', 'None', 'null', 'NULL'):
                return v
        return None

    category_id = first_non_empty('category_id')
    category_text = first_non_empty('category')
    client_id = first_non_empty('client_id') or first_non_empty('client')
    organization_id = first_non_empty('organization_id') or first_non_empty('organization')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    qs = Ticket.objects.select_related('category', 'client', 'organization', 'status', 'assigned_to')
    if category_id and str(category_id).isdigit():
        cid = int(category_id)
        qs = qs.filter(Q(category_id=cid) | Q(category__parent_id=cid))
    elif category_text:
        qs = qs.filter(Q(category__name__icontains=category_text) | Q(category__parent__name__icontains=category_text))
    if client_id and str(client_id).isdigit():
        qs = qs.filter(client_id=int(client_id))
    if organization_id and str(organization_id).isdigit():
        qs = qs.filter(organization_id=int(organization_id))
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Формируем XLSX
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Обращения'
    headers = ['ID', 'Заголовок', 'Клиент', 'Контактное лицо', 'Организация', 'Категория', 'Статус', 'Исполнитель', 'Создано', 'Дата выполнения']
    ws.append(headers)
    for t in qs.order_by('-created_at'):
        ws.append([
            t.id,
            t.title,
            t.client.name if t.client else '',
            t.client.contact_person if t.client else '',
            t.organization.name if t.organization else '',
            t.category.name if t.category else '',
            t.status.name if t.status else '',
            (t.assigned_to.get_full_name() or t.assigned_to.username) if t.assigned_to else '',
            t.created_at.strftime('%d.%m.%Y %H:%M'),
            t.resolved_at.strftime('%d.%m.%Y %H:%M') if t.resolved_at else '',
        ])

    # Автоширина
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="analytics_export_{timestamp}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def stream(request):
    """Поток сообщений Telegram"""
    qs = TelegramMessage.objects.select_related('linked_ticket').order_by('-message_date', '-id')

    # Фильтры
    group_id = request.GET.get('group_id') or request.GET.get('group')
    q = request.GET.get('q')
    if group_id and group_id not in (None, '', 'None', 'null', 'NULL'):
        qs = qs.filter(chat_id=group_id)
    if q:
        qs = qs.filter(Q(text__icontains=q) | Q(from_username__icontains=q) | Q(from_fullname__icontains=q))

    # Действие: создать тикет из сообщения
    if request.method == 'POST' and request.POST.get('action') == 'create_ticket':
        msg_id = request.POST.get('message_id')
        msg = get_object_or_404(TelegramMessage, id=msg_id)

        # Получаем данные из формы модального окна
        title = request.POST.get('title', '').strip()
        client_id = request.POST.get('client_id')
        organization_id = request.POST.get('organization_id')
        category_id = request.POST.get('category_id')

        # Поиск клиента
        client = None
        if client_id:
            try:
                client = Client.objects.get(id=client_id)
            except Client.DoesNotExist:
                pass
        
        # Если клиент не найден, ищем по from_user_id или создаем неизвестного
        if not client:
            client = Client.objects.filter(external_id=msg.from_user_id).first()
            if not client:
                client = Client.objects.filter(name='Неизвестный клиент').first()
                if not client:
                    client = Client.objects.create(name='Неизвестный клиент')

        # Поиск организации
        organization = None
        if organization_id:
            try:
                organization = Organization.objects.get(id=organization_id)
            except Organization.DoesNotExist:
                pass

        # Поиск маршрута для этой группы, клиента и организации
        route = None
        try:
            from .models import TelegramGroup
            telegram_group = TelegramGroup.objects.filter(chat_id=msg.chat_id).first()
            route = TelegramRoute.find_route(telegram_group=telegram_group, client=client, organization=organization)
        except:
            pass

        # Определяем категорию и приоритет
        category = None
        priority = 'normal'
        
        if route:
            # Используем настройки из маршрута
            category = route.category
            priority = route.priority
            
            # Формируем заголовок по шаблону маршрута
            client_name = client.name if client else 'Неизвестный клиент'
            group_name = msg.chat_title or 'Неизвестная группа'
            
            if not title:  # Если заголовок не задан вручную
                title = route.format_title(
                    group_name=group_name,
                    client_name=client_name,
                    message_id=msg.message_id,
                    message_date=msg.message_date
                )
        else:
            # Используем логику по умолчанию
            if category_id:
                try:
                    category = Category.objects.get(id=category_id)
                except Category.DoesNotExist:
                    pass
            
            if not category:
                category = Category.objects.filter(name__icontains='Обращения от поставщиков', parent__isnull=True).first() or Category.objects.first()
            
            if not title:
                title = msg.text[:100] if msg.text else 'Сообщение из Telegram'

        # Статус по умолчанию
        status = TicketStatus.objects.filter(is_final=False).order_by('order').first() or TicketStatus.objects.first()

        ticket = Ticket(
            title=title,
            description=msg.text,
            category=category,
            client=client,
            organization=organization,
            status=status,
            priority=priority,
            created_by=request.user,
        )
        ticket.external_message_id = msg.message_id
        ticket.telegram_chat_id = msg.chat_id
        ticket.telegram_chat_title = msg.chat_title
        ticket.created_at = msg.message_date
        ticket.save()

        msg.linked_ticket = ticket
        msg.linked_action = 'create_ticket'
        msg.processed_at = timezone.now()
        msg.save(update_fields=['linked_ticket', 'linked_action', 'processed_at'])

        # Добавляем информацию о маршруте в сообщение
        route_info = f' (через маршрут "{route.name}")' if route else ''
        messages.success(request, mark_safe(f'Создано обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a>{route_info}'))
        return redirect('tickets:stream')

    # Действие: решить обращение по сообщению
    if request.method == 'POST' and request.POST.get('action') == 'resolve_ticket':
        msg_id = request.POST.get('message_id')
        ticket_id = request.POST.get('ticket_id')
        msg = get_object_or_404(TelegramMessage, id=msg_id)
        ticket = get_object_or_404(Ticket, id=ticket_id)

        # Находим финальный/"Решено" статус
        resolved_status = TicketStatus.objects.filter(name='Решено').first() or TicketStatus.objects.filter(is_final=True).first()
        if not resolved_status:
            messages.error(request, 'Статус "Решено" не найден')
            return redirect('tickets:stream')

        ticket.status = resolved_status
        ticket.resolution = (msg.text or '')
        ticket.resolved_at = msg.message_date
        
        # Если taken_at пустое, устанавливаем на 1 секунду раньше resolved_at
        if not ticket.taken_at:
            ticket.taken_at = ticket.resolved_at - timezone.timedelta(seconds=1)
        
        # Если исполнитель не назначен, назначаем автора сообщения или текущего пользователя
        if not ticket.assigned_to:
            from .models import UserTelegramAccess
            uta = UserTelegramAccess.objects.select_related('user').filter(telegram_user_id=msg.from_user_id, is_allowed=True).first()
            if uta:
                ticket.assigned_to = uta.user
            else:
                ticket.assigned_to = request.user
        
        ticket.save()

        TicketAudit.objects.create(
            ticket=ticket,
            action='resolved',
            user=request.user,
            comment=f'Решено из потока: {msg.text[:50]}...'
        )

        msg.linked_ticket = ticket
        msg.linked_action = 'resolve_ticket'
        msg.processed_at = timezone.now()
        msg.save(update_fields=['linked_ticket', 'linked_action', 'processed_at'])

        messages.success(request, mark_safe(f'Обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a> переведено в Решено'))
        return redirect('tickets:stream')

    # Действие: добавить комментарий в обращение по сообщению
    if request.method == 'POST' and request.POST.get('action') == 'add_comment':
        msg_id = request.POST.get('message_id')
        ticket_id = request.POST.get('ticket_id')
        comment_text = request.POST.get('comment_text', '')
        is_internal = request.POST.get('is_internal') == 'on'
        reply_in_chat = request.POST.get('reply_in_chat') == '1'
        
        if not ticket_id or not ticket_id.isdigit():
            messages.error(request, 'Укажите корректный ID тикета')
            return redirect('tickets:stream')
        
        if not comment_text:
            messages.error(request, 'Введите текст комментария')
            return redirect('tickets:stream')
            
        msg = get_object_or_404(TelegramMessage, id=msg_id)
        ticket = get_object_or_404(Ticket, id=int(ticket_id))

        # Определяем автора
        author_type = 'user'
        author = None
        author_client = None
        # Если есть маппинг на пользователя системы
        from .models import UserTelegramAccess  # локальный импорт во избежание циклов
        uta = UserTelegramAccess.objects.select_related('user').filter(telegram_user_id=msg.from_user_id, is_allowed=True).first()
        if uta:
            author_type = 'user'
            author = uta.user
        else:
            # Если есть клиент по external_id
            cl = Client.objects.filter(external_id=msg.from_user_id).first()
            if cl:
                author_type = 'client'
                author_client = cl
            else:
                # Иначе "Неизвестный клиент"
                author_type = 'client'
                author_client = Client.objects.filter(name='Неизвестный клиент').first()
                if not author_client:
                    author_client = Client.objects.create(name='Неизвестный клиент')

        comment = TicketComment(
            ticket=ticket,
            content=comment_text,
            is_internal=is_internal,
            created_at=timezone.now(),
            author_type='user',
            author=request.user,
        )
        comment.save()

        # Обновляем сообщение в потоке
        msg.linked_ticket = ticket
        msg.linked_action = 'add_comment'
        msg.processed_at = timezone.now()
        msg.save()

        # Если запрошена отправка в Telegram
        if reply_in_chat and ticket.telegram_chat_id and ticket.external_message_id:
            try:
                import logging
                import asyncio
                from telegram.ext import Application
                from django.conf import settings
                
                logger = logging.getLogger(__name__)
                logger.info(f"Attempting to send Telegram comment: chat_id={ticket.telegram_chat_id}, message_id={msg.message_id}")
                
                bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                if bot_token:
                    # Создаем асинхронную функцию для отправки комментария
                    async def send_telegram_comment():
                        application = Application.builder().token(bot_token).build()
                        result = await application.bot.send_message(
                            chat_id=ticket.telegram_chat_id,
                            text=comment_text,
                            reply_to_message_id=int(msg.message_id)
                        )
                        return result
                    
                    # Запускаем асинхронную функцию
                    result = asyncio.run(send_telegram_comment())
                    logger.info(f"Telegram comment sent successfully: {result.message_id}")
                    
                    # Добавляем отправленное сообщение в поток
                    try:
                        TelegramMessage.objects.create(
                            message_id=str(result.message_id),
                            chat_id=str(ticket.telegram_chat_id),
                            chat_title=ticket.telegram_chat_title or '',
                            from_user_id=str(request.user.id),
                            from_username=request.user.username,
                            from_fullname=request.user.get_full_name() or request.user.username,
                            text=comment_text,
                            message_date=timezone.now(),
                            created_at=timezone.now(),
                            linked_ticket=ticket,
                            linked_action='add_comment',
                            reply_to_message_id=str(msg.message_id)
                        )
                        logger.info(f"Added comment message to stream: {result.message_id}")
                    except Exception as stream_error:
                        logger.error(f"Failed to add comment message to stream: {stream_error}", exc_info=True)
                    
                    messages.success(request, mark_safe(f'Комментарий добавлен и отправлен в Telegram в обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a>'))
                else:
                    logger.error("TELEGRAM_BOT_TOKEN not configured")
                    messages.success(request, mark_safe(f'Комментарий добавлен в обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a> (Telegram бот не настроен)'))
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to send Telegram comment: {e}", exc_info=True)
                messages.success(request, mark_safe(f'Комментарий добавлен в обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a> (не удалось отправить в Telegram)'))
        else:
            messages.success(request, mark_safe(f'Комментарий добавлен в обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a>'))

        TicketAudit.objects.create(
            ticket=ticket,
            action='comment_added',
            user=request.user,
            comment=f'Комментарий добавлен: {comment.content[:50]}...'
        )

        return redirect('tickets:stream')

    # Массовый комментарий по выбранным сообщениям
    if request.method == 'POST' and request.POST.get('action') == 'bulk_comment':
        ticket_id = request.POST.get('ticket_id')
        ids = request.POST.getlist('selected')
        if not ticket_id or not ticket_id.isdigit():
            messages.error(request, 'Укажите корректный ID тикета для массового комментария')
            return redirect('tickets:stream')
        ticket = get_object_or_404(Ticket, id=int(ticket_id))
        msgs = TelegramMessage.objects.filter(id__in=ids).order_by('message_date')
        created = 0
        for msg in msgs:
            # Получаем флаг "внутренний" для конкретного сообщения
            is_internal = request.POST.get(f'is_internal_{msg.id}') == 'on'
            author_type = 'user'
            author = None
            author_client = None
            from .models import UserTelegramAccess
            uta = UserTelegramAccess.objects.select_related('user').filter(telegram_user_id=msg.from_user_id, is_allowed=True).first()
            if uta:
                author_type = 'user'
                author = uta.user
            else:
                cl = Client.objects.filter(external_id=msg.from_user_id).first()
                if cl:
                    author_type = 'client'
                    author_client = cl
                else:
                    author_type = 'client'
                    author_client = Client.objects.filter(name='Неизвестный клиент').first()
                    if not author_client:
                        author_client = Client.objects.create(name='Неизвестный клиент')

            TicketComment.objects.create(
                ticket=ticket,
                content=msg.text or '',
                is_internal=is_internal,
                created_at=msg.message_date,
                author_type=author_type,
                author=author,
                author_client=author_client,
                telegram_message_id=msg.message_id,  # Сохраняем ID сообщения для связи
            )
            
            # Обновляем сообщение в потоке
            msg.linked_ticket = ticket
            msg.linked_action = 'add_comment'
            msg.processed_at = timezone.now()
            msg.save()
            
            created += 1
        messages.success(request, mark_safe(f'Добавлено комментариев: {created} в обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a>'))
        return redirect('tickets:stream')

    # Массовое удаление выбранных сообщений
    if request.method == 'POST' and request.POST.get('action') == 'bulk_delete':
        ids = request.POST.getlist('selected')
        deleted, _ = TelegramMessage.objects.filter(id__in=ids).delete()
        messages.success(request, f'Удалено записей: {deleted}')
        return redirect('tickets:stream')

    # Очистка за период
    if request.method == 'POST' and request.POST.get('action') == 'cleanup_period':
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        include_processed = request.POST.get('include_processed') == 'on'
        
        if not date_from or not date_to:
            messages.error(request, 'Укажите период для очистки')
            return redirect('tickets:stream')
        
        # Базовый фильтр по датам
        qs_to_delete = TelegramMessage.objects.filter(
            message_date__date__gte=date_from, 
            message_date__date__lte=date_to
        )
        
        # Если галка "Даже обработанные" не нажата, удаляем только необработанные
        if not include_processed:
            qs_to_delete = qs_to_delete.filter(linked_ticket__isnull=True)
        
        cnt, _ = qs_to_delete.delete()
        messages.success(request, f'Удалено записей из потока: {cnt}')
        return redirect('tickets:stream')

    # Пагинация с поддержкой per_page
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100, 200]:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25
    
    paginator = Paginator(qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Маппинг для отображения на странице (только текущая страница)
    from_ids = {m.from_user_id for m in page_obj.object_list if m.from_user_id}
    clients_map = {c.external_id: c for c in Client.objects.select_related('organization').filter(external_id__in=from_ids)}
    from .models import UserTelegramAccess
    uta_map = {}
    for access in UserTelegramAccess.objects.select_related('user').filter(telegram_user_id__in=from_ids, is_allowed=True):
        uta_map[access.telegram_user_id] = access.user

    # Загружаем сообщения-ответы для отображения цитат
    reply_to_ids = {m.reply_to_message_id for m in page_obj.object_list if m.reply_to_message_id}
    reply_messages_map = {}
    if reply_to_ids:
        reply_messages = TelegramMessage.objects.filter(
            message_id__in=reply_to_ids,
            chat_id__in={m.chat_id for m in page_obj.object_list}
        ).values('message_id', 'chat_id', 'text', 'from_username', 'from_fullname', 'from_user_id')
        
        for reply in reply_messages:
            key = f"{reply['chat_id']}_{reply['message_id']}"
            reply_messages_map[key] = {
                'text': reply['text'],
                'author': reply['from_username'] or reply['from_fullname'] or 'Неизвестный',
                'from_user_id': reply['from_user_id']
            }

    # Получаем название группы для отображения в фильтре
    group_name = ''
    if group_id:
        from .models import TelegramGroup
        group = TelegramGroup.objects.filter(chat_id=group_id).first()
        if group:
            group_name = group.title or group.chat_id

    # Данные для предзаполнения модального окна создания обращения
    default_category = Category.objects.filter(name__icontains='Обращения от поставщиков', parent__isnull=True).first() or Category.objects.first()
    unknown_client = Client.objects.filter(name='Неизвестный клиент').first()
    
    # Загружаем активные маршруты для предзаполнения
    active_routes = {}
    try:
        from .models import TelegramGroup
        routes = TelegramRoute.objects.filter(is_active=True).select_related('telegram_group', 'category', 'client', 'organization')
        for route in routes:
            # Создаем ключ для маршрута на основе условий
            route_key = f"{route.telegram_group.chat_id if route.telegram_group else 'no_group'}|{route.client.id if route.client else 'no_client'}|{route.organization.id if route.organization else 'no_org'}"
            active_routes[route_key] = {
                'id': route.id,
                'name': route.name,
                'title_template': route.title_template,
                'category_id': route.category.id,
                'category_name': route.category.name,
                'priority': route.priority,
                'telegram_group_id': route.telegram_group.chat_id if route.telegram_group else None,
                'client_id': route.client.id if route.client else None,
                'organization_id': route.organization.id if route.organization else None,
            }
    except:
        pass
    
    context = {
        'page_obj': page_obj,
        'filters': {
            'group_id': group_id or '',
            'group_name': group_name,
            'q': q or '',
            'per_page': per_page,
        },
        'clients_map': clients_map,
        'uta_map': uta_map,
        'reply_messages_map': reply_messages_map,
        'default_category': default_category,
        'unknown_client': unknown_client,
        'active_routes': active_routes,
        'clients_map_json': {str(k): {'id': v.id, 'name': v.name} for k, v in clients_map.items()},
    }

    # Действие: перевести обращение в работу
    if request.method == 'POST' and request.POST.get('action') == 'set_working':
        msg_id = request.POST.get('message_id')
        ticket_id = request.POST.get('ticket_id')
        comment = request.POST.get('comment', '')
        is_internal = request.POST.get('is_internal_working') == 'on'
        
        if not ticket_id or not ticket_id.isdigit():
            messages.error(request, 'Укажите корректный ID обращения')
            return redirect('tickets:stream')
        
        try:
            ticket = Ticket.objects.get(id=int(ticket_id))
            msg = get_object_or_404(TelegramMessage, id=msg_id)
        except Ticket.DoesNotExist:
            messages.error(request, 'Обращение не найдено')
            return redirect('tickets:stream')
        
        # Проверяем, что статус можно изменить на "В работе"
        working_statuses = ['Новое', 'Ожидает ответа', 'Решено']
        if ticket.status.name not in working_statuses:
            messages.error(request, f'Нельзя перевести в работу обращение со статусом "{ticket.status.name}"')
            return redirect('tickets:stream')
        
        # Получаем статус "В работе"
        try:
            working_status = TicketStatus.objects.get(name='В работе')
        except TicketStatus.DoesNotExist:
            messages.error(request, 'Статус "В работе" не найден в системе')
            return redirect('tickets:stream')
        
        # Меняем статус
        old_status = ticket.status
        ticket.status = working_status
        ticket.assigned_to = request.user
        
        # Устанавливаем время взятия в работу, если еще не установлено
        if not ticket.taken_at:
            ticket.taken_at = timezone.now()
        
        ticket.save()
        
        # Создаем комментарий с текстом сообщения из потока
        TicketComment.objects.create(
            ticket=ticket,
            author=request.user,
            author_type='user',
            content=msg.text or '',
            is_internal=is_internal,
            telegram_message_id=msg.message_id,
            created_at=msg.message_date
        )
        
        # Создаем комментарий с пользовательским текстом (если есть)
        if comment:
            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                author_type='user',
                content=comment,
                is_internal=is_internal,
                telegram_message_id=None,  # Не связываем с Telegram сообщением
                created_at=timezone.now()
            )
        
        # Создаем внутренний комментарий о смене статуса
        TicketComment.objects.create(
            ticket=ticket,
            author=request.user,
            author_type='user',
            content=f'Статус изменен с "{old_status.name}" на "В работу"',
            is_internal=True,  # Всегда внутренний
            telegram_message_id=None,  # Не связываем с Telegram сообщением
            created_at=msg.message_date  # Используем время сообщения из потока
        )
        
        # Обновляем сообщение в потоке
        msg.linked_ticket = ticket
        msg.linked_action = 'set_working'
        msg.processed_at = timezone.now()
        msg.save()
        
        # Создаем запись аудита
        audit_comment = f'Статус изменен на "В работу"'
        if comment:
            audit_comment += f': {comment[:50]}...'
        TicketAudit.objects.create(
            ticket=ticket,
            action='status_changed',
            user=request.user,
            comment=audit_comment
        )
        
        messages.success(request, mark_safe(f'Обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a> переведено в работу'))
        return redirect('tickets:stream')

    # Действие: перевести обращение в ожидание
    if request.method == 'POST' and request.POST.get('action') == 'set_waiting':
        msg_id = request.POST.get('message_id')
        ticket_id = request.POST.get('ticket_id')
        comment = request.POST.get('comment', '')
        is_internal = request.POST.get('is_internal_waiting') == 'on'
        
        if not ticket_id or not ticket_id.isdigit():
            messages.error(request, 'Укажите корректный ID обращения')
            return redirect('tickets:stream')
        
        try:
            ticket = Ticket.objects.get(id=int(ticket_id))
            msg = get_object_or_404(TelegramMessage, id=msg_id)
        except Ticket.DoesNotExist:
            messages.error(request, 'Обращение не найдено')
            return redirect('tickets:stream')
        
        # Проверяем, что статус можно изменить на "Ожидает ответа"
        waiting_statuses = ['В работе', 'Новое']
        if ticket.status.name not in waiting_statuses:
            messages.error(request, f'Нельзя перевести в ожидание обращение со статусом "{ticket.status.name}"')
            return redirect('tickets:stream')
        
        # Получаем статус "Ожидает ответа"
        try:
            waiting_status = TicketStatus.objects.get(name='Ожидает ответа')
        except TicketStatus.DoesNotExist:
            messages.error(request, 'Статус "Ожидает ответа" не найден в системе')
            return redirect('tickets:stream')
        
        # Меняем статус
        old_status = ticket.status
        ticket.status = waiting_status
        
        # Если переводим из "Новое" в "Ожидает ответа", устанавливаем время взятия в работу
        if old_status.name == 'Новое':
            ticket.taken_at = timezone.now()
        
        ticket.save()
        
        # Создаем комментарий с текстом сообщения из потока
        TicketComment.objects.create(
            ticket=ticket,
            author=request.user,
            author_type='user',
            content=msg.text or '',
            is_internal=is_internal,
            telegram_message_id=msg.message_id,
            created_at=msg.message_date
        )
        
        # Создаем комментарий с пользовательским текстом (если есть)
        if comment:
            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                author_type='user',
                content=comment,
                is_internal=is_internal,
                telegram_message_id=None,  # Не связываем с Telegram сообщением
                created_at=timezone.now()
            )
        
        # Создаем внутренний комментарий о смене статуса
        TicketComment.objects.create(
            ticket=ticket,
            author=request.user,
            author_type='user',
            content=f'Статус изменен с "{old_status.name}" на "Ожидает ответа"',
            is_internal=True,  # Всегда внутренний
            telegram_message_id=None,  # Не связываем с Telegram сообщением
            created_at=msg.message_date  # Используем время сообщения из потока
        )
        
        # Обновляем сообщение в потоке
        msg.linked_ticket = ticket
        msg.linked_action = 'set_waiting'
        msg.processed_at = timezone.now()
        msg.save()
        
        # Создаем запись аудита
        audit_comment = f'Статус изменен на "Ожидает ответа"'
        if comment:
            audit_comment += f': {comment[:50]}...'
        TicketAudit.objects.create(
            ticket=ticket,
            action='status_changed',
            user=request.user,
            comment=audit_comment
        )
        
        messages.success(request, mark_safe(f'Обращение <a href="{reverse("tickets:ticket_detail", args=[ticket.id])}" target="_blank">#{ticket.id}</a> переведено в ожидание'))
        return redirect('tickets:stream')

    return render(request, 'tickets/stream.html', context)


@login_required
def get_active_tickets(request):
    """API: получить все активные обращения для выбора (только для решения)"""
    query = request.GET.get('q', '')
    tickets = Ticket.objects.filter(
        status__is_final=False
    ).select_related('client', 'status', 'category').order_by('-created_at')
    
    if query:
        tickets = tickets.filter(
            Q(id__icontains=query) |
            Q(title__icontains=query) |
            Q(client__name__icontains=query) |
            Q(category__name__icontains=query)
        )
    
    # Ограничиваем количество результатов
    tickets = tickets[:50]
    
    results = []
    for ticket in tickets:
        results.append({
            'id': ticket.id,
            'title': ticket.title,
            'client_name': ticket.client.name,
            'status_name': ticket.status.name,
            'category_name': ticket.category.name,
            'created_at': timezone.localtime(ticket.created_at).strftime('%d.%m.%Y %H:%M'),
        })
    
    return JsonResponse({'results': results})


@login_required
def get_all_tickets(request):
    """API: получить все обращения для выбора (для комментариев)"""
    query = request.GET.get('q', '')
    tickets = Ticket.objects.all().select_related('client', 'status', 'category').order_by('-created_at')
    
    if query:
        tickets = tickets.filter(
            Q(id__icontains=query) |
            Q(title__icontains=query) |
            Q(client__name__icontains=query) |
            Q(category__name__icontains=query)
        )
    
    # Ограничиваем количество результатов
    tickets = tickets[:50]
    
    results = []
    for ticket in tickets:
        results.append({
            'id': ticket.id,
            'title': ticket.title,
            'client_name': ticket.client.name,
            'status_name': ticket.status.name,
            'created_at': timezone.localtime(ticket.created_at).strftime('%d.%m.%Y %H:%M'),
        })
    
    return JsonResponse({'results': results})


@login_required
def get_unresolved_tickets(request):
    """API: получить все нерешенные обращения для выбора"""
    query = request.GET.get('q', '')
    tickets = Ticket.objects.filter(
        status__is_final=False
    ).select_related('client', 'status', 'category').order_by('-created_at')
    
    if query:
        tickets = tickets.filter(
            Q(id__icontains=query) |
            Q(title__icontains=query) |
            Q(client__name__icontains=query) |
            Q(category__name__icontains=query)
        )
    
    # Ограничиваем количество результатов
    tickets = tickets[:50]
    
    results = []
    for ticket in tickets:
        results.append({
            'id': ticket.id,
            'title': ticket.title,
            'client_name': ticket.client.name,
            'status_name': ticket.status.name,
            'category_name': ticket.category.name,
            'created_at': timezone.localtime(ticket.created_at).strftime('%d.%m.%Y %H:%M'),
        })
    
    return JsonResponse({'results': results})


@login_required
def get_working_tickets(request):
    """API: получить обращения для перевода в работу (Новое, Ожидает ответа, Решено)"""
    query = request.GET.get('q', '')
    
    # Получаем статусы для перевода в работу
    working_statuses = ['Новое', 'Ожидает ответа', 'Решено']
    
    tickets = Ticket.objects.filter(
        status__name__in=working_statuses
    ).select_related('client', 'category', 'status', 'assigned_to').order_by('-created_at')
    
    if query:
        tickets = tickets.filter(
            Q(id__icontains=query) |
            Q(title__iregex=query) |
            Q(client__name__iregex=query) |
            Q(category__name__iregex=query)
        )
    
    results = []
    for ticket in tickets:
        results.append({
            'id': ticket.id,
            'text': f"#{ticket.id} - {ticket.title} ({ticket.status.name})",
            'title': ticket.title,
            'status_name': ticket.status.name,
            'client_name': ticket.client.name if ticket.client else 'Не указан',
            'category_name': ticket.category.name if ticket.category else 'Не указана',
            'created_at': timezone.localtime(ticket.created_at).strftime('%d.%m.%Y %H:%M'),
        })
    
    return JsonResponse({'results': results})


@login_required
def get_waiting_tickets(request):
    """API: получить обращения для перевода в ожидание (В работе, Новое)"""
    query = request.GET.get('q', '')
    
    # Получаем статусы для перевода в ожидание
    waiting_statuses = ['В работе', 'Новое']
    
    tickets = Ticket.objects.filter(
        status__name__in=waiting_statuses
    ).select_related('client', 'category', 'status', 'assigned_to').order_by('-created_at')
    
    if query:
        tickets = tickets.filter(
            Q(id__icontains=query) |
            Q(title__iregex=query) |
            Q(client__name__iregex=query) |
            Q(category__name__iregex=query)
        )
    
    results = []
    for ticket in tickets:
        results.append({
            'id': ticket.id,
            'text': f"#{ticket.id} - {ticket.title} ({ticket.status.name})",
            'title': ticket.title,
            'status_name': ticket.status.name,
            'client_name': ticket.client.name if ticket.client else 'Не указан',
            'category_name': ticket.category.name if ticket.category else 'Не указана',
            'created_at': timezone.localtime(ticket.created_at).strftime('%d.%m.%Y %H:%M'),
        })
    
    return JsonResponse({'results': results})


# ===== ORGANIZATION VIEWS =====

@login_required
def organization_list(request):
    """Список организаций"""
    search_query = request.GET.get("search", "")
    organizations = Organization.objects.filter(is_active=True)
    
    if search_query:
        organizations = organizations.filter(
            Q(name__iregex=search_query) |
            Q(comment__iregex=search_query)
        )
    
    # Добавляем количество обращений для каждой организации
    organizations = organizations.annotate(
        ticket_count=Count("ticket")
    ).order_by("name")
    
    # Пагинация
    paginator = Paginator(organizations, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "total_count": organizations.count(),
    }
    return render(request, "tickets/organization_list.html", context)


@login_required
def organization_detail(request, organization_id):
    """Детальная страница организации"""
    organization = get_object_or_404(Organization, id=organization_id)
    
    # Получаем обращения организации
    tickets = Ticket.objects.filter(organization=organization).select_related(
        "client", "status", "category", "assigned_to"
    ).order_by("-created_at")
    
    # Статистика
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status__is_final=False).count()
    
    # Пагинация обращений
    paginator = Paginator(tickets, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    context = {
        "organization": organization,
        "page_obj": page_obj,
        "ticket_count": total_tickets,
        "total_tickets": total_tickets,
        "open_tickets": open_tickets,
    }
    return render(request, "tickets/organization_detail.html", context)


@login_required
def organization_edit(request, organization_id):
    """Редактирование организации"""
    organization = get_object_or_404(Organization, id=organization_id)
    
    if request.method == "POST":
        form = OrganizationForm(request.POST, instance=organization)
        if form.is_valid():
            form.save()
            messages.success(request, f"Организация \"{organization.name}\" успешно обновлена!")
            return redirect("tickets:organization_detail", organization_id=organization.id)
    else:
        form = OrganizationForm(instance=organization)
    
    context = {
        "form": form,
        "organization": organization,
        "is_edit": True,
    }
    return render(request, "tickets/organization_form.html", context)


@login_required
def organization_create(request):
    """Создание новой организации"""
    if request.method == 'POST':
        form = OrganizationForm(request.POST)
        if form.is_valid():
            organization = form.save()
            messages.success(request, f'Организация "{organization.name}" успешно создана!')
            return redirect('tickets:organization_detail', organization_id=organization.id)
    else:
        form = OrganizationForm()
    
    context = {
        'form': form,
        'is_edit': False,
    }
    return render(request, 'tickets/organization_form.html', context)
